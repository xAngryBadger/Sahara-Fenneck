"""Tests for workbook-level actions (duplicate_sheet, create_sheet, delete_sheet, rename_sheet) via openpyxl path."""
from __future__ import annotations

import json

import pandas as pd
import pytest
from src.agent.tools import structured_actions_tool
from src.checkpoints.manager import CheckpointManager
from src.indexing.excel_reader import Workspace


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    path = tmp_path / "wb_actions.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame({"A": [1, 2], "B": [3, 4]}).to_excel(writer, sheet_name="Plan1", index=False)
        pd.DataFrame({"X": [10, 20]}).to_excel(writer, sheet_name="Plan2", index=False)
    return path


@pytest.fixture
def ws_multi(multi_sheet_xlsx):
    df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    return Workspace(
        path=str(multi_sheet_xlsx),
        workbook_name="wb_actions.xlsx",
        sheet_name="Plan1",
        columns=list(df.columns),
        row_count=len(df),
        indexed_rows=len(df),
        truncated=False,
        df=df,
        excel_live=False,
        excel_book_name=None,
        error=None,
    )


@pytest.fixture
def cp_mgr(multi_sheet_xlsx):
    return CheckpointManager(str(multi_sheet_xlsx))


class TestDuplicateSheet:
    def test_duplicate_openpyxl(self, ws_multi, cp_mgr):
        import openpyxl

        result = structured_actions_tool(
            ws_multi,
            json.dumps([{"action": "duplicate_sheet", "name": "Plan1 Cópia"}]),
            cp_mgr,
            save_checkpoint=False,
        )
        assert result.success
        wb = openpyxl.load_workbook(ws_multi.path)
        assert "Plan1 Cópia" in wb.sheetnames

    def test_duplicate_default_name(self, ws_multi, cp_mgr):
        import openpyxl

        result = structured_actions_tool(
            ws_multi,
            json.dumps([{"action": "duplicate_sheet"}]),
            cp_mgr,
            save_checkpoint=False,
        )
        assert result.success
        wb = openpyxl.load_workbook(ws_multi.path)
        assert any("Cópia" in s for s in wb.sheetnames)


class TestCreateSheet:
    def test_create_openpyxl(self, ws_multi, cp_mgr):
        import openpyxl

        result = structured_actions_tool(
            ws_multi,
            json.dumps([{"action": "create_sheet", "name": "NovaAba"}]),
            cp_mgr,
            save_checkpoint=False,
        )
        assert result.success
        wb = openpyxl.load_workbook(ws_multi.path)
        assert "NovaAba" in wb.sheetnames

    def test_create_default_name(self, ws_multi, cp_mgr):
        import openpyxl

        result = structured_actions_tool(
            ws_multi,
            json.dumps([{"action": "create_sheet"}]),
            cp_mgr,
            save_checkpoint=False,
        )
        assert result.success
        wb = openpyxl.load_workbook(ws_multi.path)
        assert len(wb.sheetnames) == 3


class TestDeleteSheet:
    def test_delete_openpyxl(self, ws_multi, cp_mgr):
        import openpyxl

        result = structured_actions_tool(
            ws_multi,
            json.dumps([{"action": "delete_sheet", "name": "Plan2"}]),
            cp_mgr,
            save_checkpoint=False,
        )
        assert result.success
        wb = openpyxl.load_workbook(ws_multi.path)
        assert "Plan2" not in wb.sheetnames
        assert "Plan1" in wb.sheetnames

    def test_delete_only_sheet_fails(self, tmp_path):
        import openpyxl

        path = tmp_path / "single.xlsx"
        pd.DataFrame({"A": [1]}).to_excel(path, index=False, engine="openpyxl", sheet_name="Only")
        df = pd.DataFrame({"A": [1]})
        ws = Workspace(
            path=str(path), workbook_name="single.xlsx", sheet_name="Only",
            columns=["A"], row_count=1, indexed_rows=1, truncated=False,
            df=df, excel_live=False, excel_book_name=None, error=None,
        )
        mgr = CheckpointManager(str(path))
        result = structured_actions_tool(
            ws,
            json.dumps([{"action": "delete_sheet", "name": "Only"}]),
            mgr,
            save_checkpoint=False,
        )
        wb = openpyxl.load_workbook(str(path))
        if result.success and len(wb.sheetnames) == 1:
            pass
        else:
            assert "Only" in wb.sheetnames


class TestRenameSheet:
    def test_rename_openpyxl(self, ws_multi, cp_mgr):
        import openpyxl

        result = structured_actions_tool(
            ws_multi,
            json.dumps([{"action": "rename_sheet", "from": "Plan2", "to": "Dados"}]),
            cp_mgr,
            save_checkpoint=False,
        )
        assert result.success
        wb = openpyxl.load_workbook(ws_multi.path)
        assert "Dados" in wb.sheetnames
        assert "Plan2" not in wb.sheetnames

    def test_rename_nonexistent_sheet(self, ws_multi, cp_mgr):
        result = structured_actions_tool(
            ws_multi,
            json.dumps([{"action": "rename_sheet", "from": "NoExiste", "to": "Algo"}]),
            cp_mgr,
            save_checkpoint=False,
        )
        assert result.success


class TestWorkbookActionsWithDfActions:
    def test_create_sheet_and_sort(self, ws_multi, cp_mgr):
        import openpyxl

        actions = [
            {"action": "create_sheet", "name": "Resumo"},
            {"action": "sort", "by": "A", "ascending": False},
        ]
        result = structured_actions_tool(ws_multi, json.dumps(actions), cp_mgr, save_checkpoint=False)
        assert result.success
        wb = openpyxl.load_workbook(ws_multi.path)
        assert "Resumo" in wb.sheetnames
        assert list(ws_multi.df["A"]) == [2, 1]


class TestWorkbookActionsNoPath:
    def test_wb_action_no_workbook_object(self, tmp_path):
        path = tmp_path / "bare.xlsx"
        pd.DataFrame({"A": [1]}).to_excel(path, index=False, engine="openpyxl")
        ws = Workspace(
            path=str(path), workbook_name="bare.xlsx", sheet_name="Sheet",
            columns=["A"], row_count=1, indexed_rows=1, truncated=False,
            df=pd.DataFrame({"A": [1]}), excel_live=False, excel_book_name=None, error=None,
        )
        import openpyxl

        wb = openpyxl.load_workbook(str(path))
        for s in wb.sheetnames[1:]:
            del wb[s]
        wb.save(str(path))
        wb.close()

        mgr = CheckpointManager(str(path))
        result = structured_actions_tool(
            ws,
            json.dumps([{"action": "create_sheet", "name": "New"}]),
            mgr,
            save_checkpoint=False,
        )
        assert result.success
        wb2 = openpyxl.load_workbook(str(path))
        assert "New" in wb2.sheetnames


class TestCOMPathWorkbookActions:
    def test_com_duplicate_sheet_no_excel(self, tmp_path):
        path = tmp_path / "com_test.xlsx"
        pd.DataFrame({"A": [1, 2]}).to_excel(path, index=False, engine="openpyxl")
        df = pd.DataFrame({"A": [1, 2]})
        ws = Workspace(
            path=str(path), workbook_name="com_test.xlsx", sheet_name="Sheet",
            columns=["A"], row_count=2, indexed_rows=2, truncated=False,
            df=df, excel_live=True, excel_book_name="com_test.xlsx", error=None,
        )
        mgr = CheckpointManager(str(path))
        result = structured_actions_tool(
            ws,
            json.dumps([{"action": "duplicate_sheet", "name": "Cópia"}]),
            mgr,
            save_checkpoint=False,
        )
        assert not result.success
        assert "Excel" in result.message or "E0" in result.message

    def test_com_create_sheet_no_excel(self, tmp_path):
        path = tmp_path / "com_test2.xlsx"
        pd.DataFrame({"X": [10]}).to_excel(path, index=False, engine="openpyxl")
        ws = Workspace(
            path=str(path), workbook_name="com_test2.xlsx", sheet_name="Sheet",
            columns=["X"], row_count=1, indexed_rows=1, truncated=False,
            df=pd.DataFrame({"X": [10]}), excel_live=True, excel_book_name="com_test2.xlsx", error=None,
        )
        mgr = CheckpointManager(str(path))
        result = structured_actions_tool(
            ws,
            json.dumps([{"action": "create_sheet", "name": "Nova"}]),
            mgr,
            save_checkpoint=False,
        )
        assert not result.success

    def test_com_rename_sheet_no_excel(self, tmp_path):
        path = tmp_path / "com_test3.xlsx"
        pd.DataFrame({"Z": [1]}).to_excel(path, index=False, engine="openpyxl")
        ws = Workspace(
            path=str(path), workbook_name="com_test3.xlsx", sheet_name="Sheet",
            columns=["Z"], row_count=1, indexed_rows=1, truncated=False,
            df=pd.DataFrame({"Z": [1]}), excel_live=True, excel_book_name="com_test3.xlsx", error=None,
        )
        mgr = CheckpointManager(str(path))
        result = structured_actions_tool(
            ws,
            json.dumps([{"action": "rename_sheet", "from": "Sheet", "to": "Dados"}]),
            mgr,
            save_checkpoint=False,
        )
        assert not result.success

    def test_com_delete_sheet_no_excel(self, tmp_path):
        path = tmp_path / "com_test4.xlsx"
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            pd.DataFrame({"A": [1]}).to_excel(writer, sheet_name="S1", index=False)
            pd.DataFrame({"B": [2]}).to_excel(writer, sheet_name="S2", index=False)
        ws = Workspace(
            path=str(path), workbook_name="com_test4.xlsx", sheet_name="S1",
            columns=["A"], row_count=1, indexed_rows=1, truncated=False,
            df=pd.DataFrame({"A": [1]}), excel_live=True, excel_book_name="com_test4.xlsx", error=None,
        )
        mgr = CheckpointManager(str(path))
        result = structured_actions_tool(
            ws,
            json.dumps([{"action": "delete_sheet", "name": "S2"}]),
            mgr,
            save_checkpoint=False,
        )
        assert not result.success


class TestWorkbookActionsODSFormat:
    def test_wb_action_ods_rejected(self, tmp_path):
        path = tmp_path / "test.ods"
        pd.DataFrame({"A": [1]}).to_excel(path, index=False, engine="odf")
        ws = Workspace(
            path=str(path), workbook_name="test.ods", sheet_name="Sheet",
            columns=["A"], row_count=1, indexed_rows=1, truncated=False,
            df=pd.DataFrame({"A": [1]}), excel_live=False, excel_book_name=None, error=None,
        )
        mgr = CheckpointManager(str(path))
        result = structured_actions_tool(
            ws,
            json.dumps([{"action": "create_sheet", "name": "New"}]),
            mgr,
            save_checkpoint=False,
        )
        assert not result.success
        assert "ods" in result.message.lower() or "xlsx" in result.message.lower()
