# -*- coding: utf-8 -*-
"""
Leitura de planilhas (Excel aberto via COM e arquivos .xlsx/.xlsm/.xls).
Suporta indexação de múltiplas planilhas/abas e limites por aba para desempenho.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

try:
    import pandas as pd
except ImportError:  # pragma: no cover
    pd = None

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
DEFAULT_MAX_ROWS = 5000


def _normalize_limit(max_rows: int) -> Optional[int]:
    """Normaliza limite de linhas: <=0 significa sem limite."""
    try:
        value = int(max_rows)
    except Exception:
        value = DEFAULT_MAX_ROWS
    if value <= 0:
        return None
    return max(1, value)


@dataclass
class Workspace:
    """Representa uma aba indexada para uso do agente."""

    path: str
    workbook_name: str
    sheet_name: str
    columns: list[str]
    row_count: int  # total de linhas de dados na aba
    indexed_rows: int  # linhas efetivamente carregadas para o contexto
    truncated: bool = False
    df: Optional["pd.DataFrame"] = None
    excel_live: bool = False  # True quando veio de Excel aberto (COM)
    excel_book_name: Optional[str] = None
    error: Optional[str] = None

    def summary_one_line(self) -> str:
        if self.error:
            return self.error
        extra = " (amostra)" if self.truncated else ""
        name = self.workbook_name or Path(self.path).name
        return (
            f"{name} | aba: {self.sheet_name} | {len(self.columns)} colunas | "
            f"{self.indexed_rows}/{self.row_count} linhas{extra}"
        )


def is_excel_file(path: str) -> bool:
    return Path(path).suffix.lower() in EXCEL_EXTENSIONS


def _safe_headers(raw_headers: list[object], col_count: int) -> list[str]:
    headers = []
    for i in range(col_count):
        v = raw_headers[i] if i < len(raw_headers) else None
        text = str(v).strip() if v is not None else ""
        headers.append(text if text else f"Col{i + 1}")
    return headers


def _rows_to_df(headers: list[str], rows: list[list[object]]) -> Optional["pd.DataFrame"]:
    if pd is None:
        return None
    if not rows:
        return pd.DataFrame(columns=headers)
    return pd.DataFrame(rows, columns=headers)


def index_from_path(file_path: str, sheet_name: Optional[str] = None, max_rows: int = DEFAULT_MAX_ROWS) -> Workspace:
    """Indexa uma aba única de arquivo Excel (por padrão aba ativa/primeira)."""
    all_items = index_file_multi(file_path, include_all_sheets=True, max_rows=max_rows)
    if not all_items:
        return Workspace(
            path=str(Path(file_path).resolve()),
            workbook_name=Path(file_path).name,
            sheet_name="",
            columns=[],
            row_count=0,
            indexed_rows=0,
            error="Não foi possível indexar arquivo.",
        )
    if sheet_name:
        for item in all_items:
            if item.sheet_name == sheet_name:
                return item
    return all_items[0]


def index_file_multi(file_path: str, include_all_sheets: bool = False, max_rows: int = DEFAULT_MAX_ROWS) -> list[Workspace]:
    """Indexa um arquivo e retorna uma Workspace por aba."""
    p = Path(file_path).resolve()
    if not p.exists():
        return [
            Workspace(
                path=str(p),
                workbook_name=p.name,
                sheet_name="",
                columns=[],
                row_count=0,
                indexed_rows=0,
                error="Arquivo não encontrado.",
            )
        ]
    if not is_excel_file(str(p)):
        return [
            Workspace(
                path=str(p),
                workbook_name=p.name,
                sheet_name="",
                columns=[],
                row_count=0,
                indexed_rows=0,
                error="Formato inválido. Selecione apenas arquivos de planilha.",
            )
        ]

    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        selected_names = wb.sheetnames if include_all_sheets else [wb.active.title]
        workspaces: list[Workspace] = []

        limit = _normalize_limit(max_rows)

        for name in selected_names:
            ws = wb[name]
            max_row = int(ws.max_row or 0)
            max_col = int(ws.max_column or 0)
            if max_row <= 0 or max_col <= 0:
                workspaces.append(
                    Workspace(
                        path=str(p),
                        workbook_name=p.name,
                        sheet_name=name,
                        columns=[],
                        row_count=0,
                        indexed_rows=0,
                        error="Aba vazia.",
                    )
                )
                continue

            # Cabeçalho = primeira linha
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            raw_headers = list(first_row) if first_row else []
            headers = _safe_headers(raw_headers, max_col)

            total_rows = max(0, max_row - 1)
            take_rows = total_rows if limit is None else min(total_rows, limit)
            data_rows: list[list[object]] = []

            if take_rows > 0:
                for row in ws.iter_rows(min_row=2, max_row=take_rows + 1, max_col=max_col, values_only=True):
                    data_rows.append(list(row))

            workspaces.append(
                Workspace(
                    path=str(p),
                    workbook_name=p.name,
                    sheet_name=name,
                    columns=headers,
                    row_count=total_rows,
                    indexed_rows=len(data_rows),
                    truncated=total_rows > len(data_rows),
                    df=_rows_to_df(headers, data_rows),
                    excel_live=False,
                    excel_book_name=None,
                    error=None,
                )
            )

        wb.close()
        return workspaces

    except Exception as e:
        return [
            Workspace(
                path=str(p),
                workbook_name=p.name,
                sheet_name="",
                columns=[],
                row_count=0,
                indexed_rows=0,
                error=f"Arquivo: {e!s}",
            )
        ]


def index_from_excel(include_all_sheets: bool = False, max_rows: int = DEFAULT_MAX_ROWS) -> Workspace:
    """Indexa uma única aba do Excel aberto (por padrão, aba ativa do workbook ativo)."""
    items = index_open_excel_workbooks(include_all_sheets=include_all_sheets, max_rows=max_rows)
    if not items:
        return Workspace(
            path="",
            workbook_name="",
            sheet_name="",
            columns=[],
            row_count=0,
            indexed_rows=0,
            error="Nenhuma planilha aberta no Excel.",
        )
    return items[0]


def index_open_excel_workbooks(include_all_sheets: bool = False, max_rows: int = DEFAULT_MAX_ROWS) -> list[Workspace]:
    """
    Indexa workbooks abertos no Excel Desktop via COM.
    Se include_all_sheets=False, indexa apenas a aba ativa de cada workbook.
    """
    workspaces: list[Workspace] = []
    limit = _normalize_limit(max_rows)

    pythoncom = None
    try:
        import pythoncom  # type: ignore

        pythoncom.CoInitialize()

        import win32com.client  # type: ignore

        excel = win32com.client.GetActiveObject("Excel.Application")
        if excel is None or excel.Workbooks.Count == 0:
            return [
                Workspace(
                    path="",
                    workbook_name="",
                    sheet_name="",
                    columns=[],
                    row_count=0,
                    indexed_rows=0,
                    error="Nenhuma planilha aberta no Excel.",
                )
            ]

        for wb in excel.Workbooks:
            try:
                wb_name = str(wb.Name)
                wb_path = str(wb.FullName) if getattr(wb, "FullName", None) else wb_name
                sheet_names = [str(s.Name) for s in wb.Worksheets] if include_all_sheets else [str(wb.ActiveSheet.Name)]

                for sname in sheet_names:
                    ws = wb.Worksheets(sname)
                    used = ws.UsedRange
                    rows_count = int(getattr(used.Rows, "Count", 0) or 0)
                    cols_count = int(getattr(used.Columns, "Count", 0) or 0)

                    if rows_count <= 0 or cols_count <= 0:
                        workspaces.append(
                            Workspace(
                                path=wb_path,
                                workbook_name=wb_name,
                                sheet_name=sname,
                                columns=[],
                                row_count=0,
                                indexed_rows=0,
                                excel_live=True,
                                excel_book_name=wb_name,
                                error="Aba vazia.",
                            )
                        )
                        continue

                    # Cabeçalho
                    raw_headers = [ws.Cells(1, c).Value for c in range(1, cols_count + 1)]
                    headers = _safe_headers(raw_headers, cols_count)

                    total_rows = max(0, rows_count - 1)
                    take_rows = total_rows if limit is None else min(total_rows, limit)
                    data_rows: list[list[object]] = []

                    for r in range(2, take_rows + 2):
                        data_rows.append([ws.Cells(r, c).Value for c in range(1, cols_count + 1)])

                    workspaces.append(
                        Workspace(
                            path=wb_path,
                            workbook_name=wb_name,
                            sheet_name=sname,
                            columns=headers,
                            row_count=total_rows,
                            indexed_rows=len(data_rows),
                            truncated=total_rows > len(data_rows),
                            df=_rows_to_df(headers, data_rows),
                            excel_live=True,
                            excel_book_name=wb_name,
                            error=None,
                        )
                    )
            except Exception as wb_err:
                workspaces.append(
                    Workspace(
                        path="",
                        workbook_name=getattr(wb, "Name", ""),
                        sheet_name="",
                        columns=[],
                        row_count=0,
                        indexed_rows=0,
                        excel_live=True,
                        excel_book_name=getattr(wb, "Name", ""),
                        error=f"Workbook com erro: {wb_err!s}",
                    )
                )

        if not workspaces:
            workspaces.append(
                Workspace(
                    path="",
                    workbook_name="",
                    sheet_name="",
                    columns=[],
                    row_count=0,
                    indexed_rows=0,
                    error="Nenhuma aba indexada do Excel.",
                )
            )

        return workspaces

    except Exception as e:
        return [
            Workspace(
                path="",
                workbook_name="",
                sheet_name="",
                columns=[],
                row_count=0,
                indexed_rows=0,
                error=f"Excel: {e!s}",
            )
        ]
    finally:
        if pythoncom is not None:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def get_workspace_summary(ws: Workspace) -> str:
    """Texto resumido da workspace para o LLM (GetData)."""
    if ws.error:
        return ws.error

    lines = [
        f"Arquivo: {ws.workbook_name}",
        f"Aba: {ws.sheet_name}",
        f"Colunas ({len(ws.columns)}): {', '.join(ws.columns[:20])}{'...' if len(ws.columns) > 20 else ''}",
        f"Linhas totais: {ws.row_count}",
        f"Linhas indexadas: {ws.indexed_rows}{' (amostra)' if ws.truncated else ''}",
    ]
    if ws.df is not None and not ws.df.empty:
        lines.append("Amostra (5 primeiras linhas):")
        lines.append(ws.df.head().to_string())
    return "\n".join(lines)


def hydrate_workspace_full(ws: Workspace) -> Workspace:
    """
    Garante que a workspace contenha todas as linhas da aba.
    Se já estiver completa, retorna sem alterações.
    """
    if ws.error:
        return ws
    if not ws.truncated:
        return ws

    # Reindexa somente a aba da workspace, sem limite.
    if ws.excel_live:
        refreshed = index_open_excel_workbooks(include_all_sheets=True, max_rows=0)
        for item in refreshed:
            if item.error:
                continue
            if (item.workbook_name == ws.workbook_name) and (item.sheet_name == ws.sheet_name):
                return item
        return ws

    refreshed_items = index_file_multi(ws.path, include_all_sheets=True, max_rows=0)
    for item in refreshed_items:
        if item.error:
            continue
        if item.sheet_name == ws.sheet_name:
            return item
    return ws
