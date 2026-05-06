# -*- coding: utf-8 -*-
"""
Loop ReAct do agente: chama o LLM com prompt + tools; interpreta resposta e executa
GetData ou Optimize; repete até resposta final. Cada Optimize gera um checkpoint.
"""
from __future__ import annotations

import json
import re
from typing import Optional, Callable

from ..indexing.excel_reader import Workspace, get_workspace_summary, hydrate_workspace_full, index_from_path
from ..checkpoints.manager import CheckpointManager
from ..integrations import handle_integration_query
from .ollama_client import OllamaClient
from .tools import optimize_tool, structured_actions_tool

# ---------------------------------------------------------------------------
# Intent classifier — keyword check to reduce spurious [ACTIONS] on read-only queries.
# ---------------------------------------------------------------------------
_MODIFY_SIGNALS = re.compile(
    r"\b(ordena|sort|filtra|filter|preenche|preencha|renomeia|renomeie|"
    r"remove|remova|apaga|apague|deleta|delete|cria|crie|adiciona|adicione|"
    r"insere|insira|substitui|substitua|troca|troque|muda|mude|altera|altere|"
    r"duplica|duplicate|limpa|limpe|formata|formate|coloca|coloque|"
    r"nova aba|nova coluna|nova planilha|nova linha|aplica|aplique|"
    r"faz isso|execute|executa)\b",
    re.IGNORECASE,
)


def _is_read_only_intent(query: str) -> bool:
    """Returns True when query contains no explicit modification keyword."""
    return not bool(_MODIFY_SIGNALS.search(query))


# ---------------------------------------------------------------------------
# Sheet-name detection — "aba Vendas" => switch workspace to that sheet.
# ---------------------------------------------------------------------------
def _detect_sheet_name(query: str, workspace: "Workspace") -> "Optional[str]":
    m = re.search(
        r"(?:aba|planilha|sheet|guia)(?:\s+(?:chamada|nomeada|nome|de nome))?\s+['\"]?([A-Za-z\u00C0-\u00FF0-9_\- ]{1,40})['\"]?",
        query,
        re.IGNORECASE,
    )
    if not m or not workspace.path:
        return None
    candidate = m.group(1).strip()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(workspace.path, read_only=True, data_only=True)
        names_lower = {n.lower(): n for n in wb.sheetnames}
        wb.close()
        return names_lower.get(candidate.lower())
    except Exception:
        return None


def _list_sheet_names(workspace: Workspace) -> list[str]:
    if workspace.path:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(workspace.path, read_only=True, data_only=True)
            names = list(wb.sheetnames)
            wb.close()
            return names
        except Exception:
            pass

    if workspace.excel_live and workspace.excel_book_name:
        pythoncom = None
        try:
            import pythoncom  # type: ignore
            import win32com.client  # type: ignore
            pythoncom.CoInitialize()
            excel = win32com.client.GetActiveObject("Excel.Application")
            for wb in excel.Workbooks:
                if str(getattr(wb, "Name", "")).lower() == workspace.excel_book_name.lower():
                    return [str(s.Name) for s in wb.Worksheets]
        except Exception:
            return []
        finally:
            if pythoncom is not None:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    return []


def _handle_sheet_query(query: str, workspace: Workspace) -> Optional[str]:
    q = (query or "").lower()
    if not any(word in q for word in ("aba", "abas", "sheet", "planilha", "guia")):
        return None

    names = _list_sheet_names(workspace)
    if not names:
        return None

    target = _detect_sheet_name(query, workspace)
    names_lower = {n.lower(): n for n in names}

    if any(word in q for word in ("quais", "listar", "liste", "nomes", "quantas")):
        return f"Este arquivo possui {len(names)} aba(s): {', '.join(names)}."

    if target and any(word in q for word in ("existe", "tem", "há", "ha")):
        if target.lower() in names_lower:
            return f"Sim. Existe uma aba chamada '{names_lower[target.lower()]}' neste arquivo."
        return f"Não. As abas disponíveis são: {', '.join(names)}."

    if "duas abas" in q or "2 abas" in q or "todas as abas" in q:
        return f"Este arquivo tem {len(names)} aba(s): {', '.join(names)}. Posso trabalhar em uma aba por vez, mas consigo alternar para qualquer uma delas quando você citar o nome exato."

    return None


def _switch_workspace_to_sheet(query: str, workspace: Workspace) -> Workspace:
    """If user references a different sheet by name, switch context to it."""
    target = _detect_sheet_name(query, workspace)
    if target and target != workspace.sheet_name and workspace.path:
        try:
            alt = index_from_path(workspace.path, sheet_name=target)
            if not alt.error:
                return alt
        except Exception:
            pass
    return workspace


SYSTEM = """Você é o Fennec, assistente de planilhas Excel. Responda SEMPRE em português.

IMPORTANTE: Use blocos de ferramenta SOMENTE quando o usuário pedir para MODIFICAR a planilha.
Para perguntas, análises, cálculos ou conversa, responda apenas em texto — sem [ACTIONS] nem [OPTIMIZE].
Exemplos que NÃO precisam de ferramenta: "o que é X?", "quantos registros?", "mostre o resumo", "qual a soma de Y?".

Para aplicar mudanças, use:
[ACTIONS]
{"actions": [{"action": "NOME", ...parâmetros...}]}
[/ACTIONS]

Ações disponíveis:
- sort: {"action":"sort","by":["Coluna"],"ascending":true}
- fillna: {"action":"fillna","column":"Col","value":0}
- replace: {"action":"replace","column":"Col","from":"X","to":"Y"}
- rename_column: {"action":"rename_column","from":"Antigo","to":"Novo"}
- drop_columns: {"action":"drop_columns","columns":["Col1"]}
- add_computed_column: {"action":"add_computed_column","new_column":"Total","operation":"sum","source_columns":["A","B"]}
- filter_equals: {"action":"filter_equals","column":"Col","value":"X"}
- duplicate_sheet: {"action":"duplicate_sheet","name":"Nome da nova aba"}
- create_sheet: {"action":"create_sheet","name":"Nome da nova aba"}
- delete_sheet: {"action":"delete_sheet","name":"Nome da aba a excluir"}
- rename_sheet: {"action":"rename_sheet","from":"NomeAtual","to":"NomeNovo"}

Use os nomes exatos das colunas fornecidas. Nunca invente colunas. Apenas uma ferramenta por resposta."""


def _extract_optimize(code_text: str) -> Optional[str]:
    m = re.search(r"\[OPTIMIZE\]\s*(.*?)\s*\[/OPTIMIZE\]", code_text, re.DOTALL | re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return None


def _extract_actions(code_text: str) -> Optional[str]:
    m = re.search(r"\[ACTIONS\]\s*(.*?)\s*\[/ACTIONS\]", code_text, re.DOTALL | re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return None


def _extract_loose_actions(code_text: str) -> Optional[str]:
    """Accept raw/fenced JSON when smaller models forget [ACTIONS] tags."""
    text = (code_text or "").strip()
    if not text:
        return None

    candidates = [text]

    fence = re.search(r"```(?:json)?\s*(.*?)\s*```", text, re.DOTALL | re.IGNORECASE)
    if fence:
        candidates.append(fence.group(1).strip())

    first_brace = text.find("{")
    last_brace = text.rfind("}")
    if first_brace != -1 and last_brace > first_brace:
        candidates.append(text[first_brace:last_brace + 1].strip())

    first_bracket = text.find("[")
    last_bracket = text.rfind("]")
    if first_bracket != -1 and last_bracket > first_bracket:
        candidates.append(text[first_bracket:last_bracket + 1].strip())

    seen: set[str] = set()
    for candidate in candidates:
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        try:
            raw = json.loads(candidate)
        except Exception:
            continue

        if isinstance(raw, dict) and isinstance(raw.get("actions"), list) and raw.get("actions"):
            return candidate
        if isinstance(raw, list) and raw and all(isinstance(a, dict) and a.get("action") for a in raw):
            return candidate

    return None


def _strip_tool_payload_from_response(response: str, payload: str, tool_name: str) -> str:
    """Removes tool payload from visible text, supporting tagged or raw JSON/code output."""
    text = response or ""

    if f"[{tool_name}]" in text:
        base = text.split(f"[{tool_name}]", 1)[0]
        return re.sub(r"\[/?(?:ACTIONS|OPTIMIZE)\]", "", base).strip()

    stripped = text
    if payload:
        stripped = stripped.replace(payload, "")
    stripped = re.sub(r"```(?:json)?\s*```", "", stripped, flags=re.DOTALL | re.IGNORECASE)
    stripped = re.sub(r"```(?:json)?\s*", "", stripped, flags=re.IGNORECASE)
    stripped = stripped.replace("```", "")
    return stripped.strip()


def _summarize_actions(actions_payload: str, workspace: Workspace) -> str:
    try:
        raw = json.loads(actions_payload)
        actions = raw.get("actions", []) if isinstance(raw, dict) else raw
    except Exception:
        return (
            f"A planilha `{workspace.workbook_name}` / aba `{workspace.sheet_name}` sera alterada.\n\n"
            "Nao foi possivel montar a previa detalhada das acoes, mas o agente tentou executar um bloco [ACTIONS]."
        )

    if not isinstance(actions, list):
        actions = []

    lines = [
        f"Planilha: {workspace.workbook_name or workspace.path}",
        f"Aba ativa: {workspace.sheet_name}",
        "",
        "Acoes propostas:",
    ]

    labels = {
        "sort": "Ordenar linhas",
        "fillna": "Preencher valores vazios",
        "replace": "Substituir valores",
        "rename_column": "Renomear coluna",
        "drop_columns": "Remover colunas",
        "add_computed_column": "Criar coluna calculada",
        "filter_equals": "Filtrar linhas por valor",
        "duplicate_sheet": "Duplicar aba",
        "create_sheet": "Criar nova aba",
        "delete_sheet": "Excluir aba",
        "rename_sheet": "Renomear aba",
    }

    for idx, action in enumerate(actions, 1):
        if not isinstance(action, dict):
            lines.append(f"{idx}. Acao invalida recebida do modelo")
            continue
        kind = str(action.get("action", "")).strip().lower()
        title = labels.get(kind, kind or "acao")
        details = []
        for key, value in action.items():
            if key == "action":
                continue
            details.append(f"{key}={value}")
        suffix = f" ({', '.join(details)})" if details else ""
        lines.append(f"{idx}. {title}{suffix}")

    lines.append("")
    lines.append("Ao confirmar, o Fennec criara um checkpoint antes da primeira alteracao.")
    return "\n".join(lines)


def _summarize_optimize(code: str, workspace: Workspace) -> str:
    snippet = code.strip()
    if len(snippet) > 700:
        snippet = snippet[:700].rstrip() + "\n..."
    return (
        f"Planilha: {workspace.workbook_name or workspace.path}\n"
        f"Aba ativa: {workspace.sheet_name}\n\n"
        "O agente quer executar uma transformacao personalizada na planilha:\n\n"
        f"{snippet}\n\n"
        "Ao confirmar, o Fennec criara um checkpoint antes da primeira alteracao."
    )


def run_agent(
    query: str,
    workspace: Workspace,
    ollama: Optional[OllamaClient] = None,
    on_message: Optional[Callable[[str], None]] = None,
    on_checkpoint: Optional[Callable[[str], None]] = None,
    on_confirm_change: Optional[Callable[[str], bool]] = None,
    max_steps: int = 5,
) -> str:
    """
    Executa o agente: envia a query ao LLM, interpreta ACTIONS/Optimize,
    executa tools (com checkpoint antes da primeira alteração) e retorna a resposta final.
    """
    direct_sheet_reply = _handle_sheet_query(query, workspace)
    if direct_sheet_reply is not None:
        if on_message:
            on_message(direct_sheet_reply)
        return direct_sheet_reply

    # Fast path: purely informational queries don't need tool scaffolding.
    # The system prompt already instructs the model, but this guard reduces
    # accidental [ACTIONS] output from smaller models on read-only questions.
    if _is_read_only_intent(query) and not workspace.error:
        # If user references a specific sheet, switch context
        switched = _switch_workspace_to_sheet(query, workspace)
        if switched is not workspace:
            workspace = switched
        client = ollama or OllamaClient()
        if client.is_available():
            data_summary = get_workspace_summary(workspace)
            ro_prompt = f"Dados da planilha:\n{data_summary}\n\nPergunta: {query}"
            response = client.generate(ro_prompt, system=SYSTEM)
            clean = re.sub(r"\[/?(?:ACTIONS|OPTIMIZE)\]", "", response).strip()
            if clean:
                if on_message:
                    on_message(clean)
                return clean

    if workspace.error:
        msg = f"Nenhuma planilha indexada: {workspace.error}"
        if on_message:
            on_message(msg)
        return msg

    # Garante visão completa da aba para reduzir perda de contexto por truncamento.
    refreshed = hydrate_workspace_full(workspace)
    if refreshed is not workspace:
        workspace.path = refreshed.path
        workspace.workbook_name = refreshed.workbook_name
        workspace.sheet_name = refreshed.sheet_name
        workspace.columns = refreshed.columns
        workspace.row_count = refreshed.row_count
        workspace.indexed_rows = refreshed.indexed_rows
        workspace.truncated = refreshed.truncated
        workspace.df = refreshed.df
        workspace.excel_live = refreshed.excel_live
        workspace.excel_book_name = refreshed.excel_book_name
        workspace.error = refreshed.error

    # Switch context if user explicitly references a different sheet
    workspace = _switch_workspace_to_sheet(query, workspace)

    # v2: algumas intencoes vao direto para integracoes externas
    integration_reply = handle_integration_query(query, workspace)
    if integration_reply is not None:
        if on_message:
            on_message(integration_reply)
        return integration_reply

    client = ollama or OllamaClient()
    if not client.is_available():
        msg = "Não consegui inicializar o Ollama automaticamente. Verifique se o Ollama está instalado e tente abrir o Fennec novamente."
        if on_message:
            on_message(msg)
        return msg

    cp = CheckpointManager(workspace.path, interaction_label="Otimização")

    # Dados iniciais para o LLM
    data_summary = get_workspace_summary(workspace)
    prompt = f"""Dados atuais da planilha:
{data_summary}

Pergunta do usuário: {query}

Responda e, se for aplicar alterações na planilha, prefira [ACTIONS]...[/ACTIONS].
Use [OPTIMIZE] apenas se a tarefa não puder ser representada pelas ações estruturadas."""

    messages = [prompt]
    final_answer = ""
    checkpoint_saved_for_order = False

    for step in range(max_steps):
        response = client.generate("\n\n".join(messages), system=SYSTEM)

        actions_payload = _extract_actions(response) or _extract_loose_actions(response)
        if actions_payload:
            # Validate actions payload before showing confirmation
            try:
                _raw = json.loads(actions_payload)
                _acts = _raw.get("actions", []) if isinstance(_raw, dict) else _raw
                if isinstance(_acts, list) and not _acts:
                    # Empty actions list — skip confirmation and tell user
                    if on_message:
                        on_message("O modelo não gerou ações válidas. Tente reformular seu pedido.")
                    messages.append(response)
                    messages.append("Resultado: A lista de ações estava vazia. Gere ações válidas.")
                    continue
            except Exception:
                pass

            if on_confirm_change:
                preview = _summarize_actions(actions_payload, workspace)
                approved = on_confirm_change(preview)
                if not approved:
                    final_answer = "Operacao cancelada. Nenhuma alteracao foi aplicada na planilha."
                    if on_message:
                        on_message(final_answer)
                    break

            save_checkpoint_now = not checkpoint_saved_for_order
            result = structured_actions_tool(
                workspace,
                actions_payload,
                cp,
                on_checkpoint_saved=on_checkpoint if save_checkpoint_now else None,
                save_checkpoint=save_checkpoint_now,
            )
            if save_checkpoint_now and not result.lower().startswith("erro ao salvar checkpoint"):
                checkpoint_saved_for_order = True

            if "sucesso" in result.lower():
                clean_text = _strip_tool_payload_from_response(response, actions_payload, "ACTIONS")
                final_answer = (clean_text + "\n\n" + result).strip()
                if on_message:
                    on_message(final_answer)
                break

            # Tool failed — let LLM retry with the error
            if on_message:
                on_message(f"Não foi possível aplicar: {result}")
            messages.append(response)
            messages.append(f"Resultado da ferramenta ACTIONS: {result}")
            continue

        code = _extract_optimize(response)
        if code:
            if on_confirm_change:
                preview = _summarize_optimize(code, workspace)
                approved = on_confirm_change(preview)
                if not approved:
                    final_answer = "Operacao cancelada. Nenhuma alteracao foi aplicada na planilha."
                    if on_message:
                        on_message(final_answer)
                    break

            save_checkpoint_now = not checkpoint_saved_for_order
            result = optimize_tool(
                workspace,
                code,
                cp,
                on_checkpoint_saved=on_checkpoint if save_checkpoint_now else None,
                save_checkpoint=save_checkpoint_now,
            )
            if save_checkpoint_now and not result.lower().startswith("erro ao salvar checkpoint"):
                checkpoint_saved_for_order = True

            if "sucesso" in result.lower():
                clean_text = re.sub(r"\[/?(?:ACTIONS|OPTIMIZE)\]", "", response.split("[OPTIMIZE]")[0]).strip()
                final_answer = (clean_text + "\n\n" + result).strip()
                if on_message:
                    on_message(final_answer)
                break

            # Tool failed — let LLM retry with the error
            if on_message:
                on_message(f"Não foi possível aplicar: {result}")
            messages.append(response)
            messages.append(f"Resultado da ferramenta Optimize: {result}")
            continue

        # No tool block: this is the final answer
        final_answer = response
        if on_message:
            on_message(response)
        break

    if not final_answer:
        final_answer = "Nenhuma resposta gerada."
        if on_message:
            on_message(final_answer)

    return final_answer
