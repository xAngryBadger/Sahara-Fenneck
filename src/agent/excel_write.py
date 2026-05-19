"""Escrita de dados no Excel — COM (tempo real) e openpyxl (arquivo .xlsx)."""
from __future__ import annotations

import logging
from collections.abc import Callable
from pathlib import Path

from ..checkpoints.manager import CheckpointManager
from ..indexing.excel_reader import Workspace
from .actions import _QUERY_KINDS, _WB_KINDS, _apply_actions_to_df, _classify_err, _normalize_actions
from .result import ErrCode, ToolResult

log = logging.getLogger(__name__)


def _excel_scalar(value):
    """Converte valores pandas/numpy para tipos seguros para COM do Excel."""
    try:
        import pandas as pd

        if pd.isna(value):
            return None
    except (ValueError, TypeError):
        pass

    try:
        import numpy as np

        if isinstance(value, np.generic):
            return value.item()
    except (ValueError, TypeError):
        pass

    return value


def _df_to_excel_matrix(df):
    rows = [tuple(str(c) for c in df.columns)]
    for row in df.itertuples(index=False, name=None):
        rows.append(tuple(_excel_scalar(v) for v in row))
    return tuple(rows)


def structured_actions_tool(
    workspace: Workspace,
    actions_payload: str,
    checkpoint_manager: CheckpointManager,
    on_checkpoint_saved: Callable[[str], None] | None = None,
    save_checkpoint: bool = True,
) -> ToolResult:
    """Executa ações estruturadas (JSON) sobre a planilha sem usar exec arbitrário."""
    actions, parse_err = _normalize_actions(actions_payload)
    if parse_err:
        return ToolResult.err(parse_err, code=ErrCode.PARSE_ACTIONS)

    checkpoint_suffix = ""
    if save_checkpoint:
        try:
            info = checkpoint_manager.save_checkpoint(workspace, label="Antes da ordem")
            if on_checkpoint_saved:
                on_checkpoint_saved(info.label)
            checkpoint_suffix = " Checkpoint salvo antes da ordem."
        except (OSError, PermissionError, RuntimeError) as cp_err:
            log.exception("Falha ao salvar checkpoint (structured_actions)")
            return ToolResult.err(f"Erro ao salvar checkpoint: {cp_err!s}", code=ErrCode.CHECKPOINT_SAVE)

    import openpyxl
    import pandas as pd

    df = workspace.df
    if df is None:
        df = pd.DataFrame(columns=workspace.columns or [])

    wb_actions = [a for a in actions if str(a.get("action", "")).strip().lower() in _WB_KINDS]
    df_actions = [a for a in actions if str(a.get("action", "")).strip().lower() not in _WB_KINDS]

    if df_actions:
        new_df, err = _apply_actions_to_df(df, df_actions)
        if err:
            if err.startswith("__REQUEST_MORE_ROWS__") or err.startswith("__ADJUST_HEADER__"):
                return ToolResult.ok(err)
            return ToolResult.err(err, code=_classify_err(err))
    else:
        new_df = df.copy() if df is not None else pd.DataFrame(columns=workspace.columns or [])

    wb_openpyxl = None
    if workspace.path and not workspace.excel_live:
        save_path = workspace.path
        original_suffix = Path(save_path).suffix.lower()
        if original_suffix in (".ods", ".xls"):
            save_path = str(Path(save_path).with_suffix(".xlsx"))
            workspace.path = save_path
            log.info("Conversão de formato: %s → %s", original_suffix, ".xlsx")
        try:
            wb_openpyxl = openpyxl.load_workbook(save_path)
        except (OSError, ValueError, KeyError):
            log.exception("Falha ao carregar workbook com openpyxl (structured_actions)")
            wb_openpyxl = None

    try:
        new_wb = wb_openpyxl
        if workspace.excel_live:
            try:
                from ..com_utils import COMContext
            except ImportError:
                COMContext = None

            if COMContext is None:
                return ToolResult.err("Excel aberto não encontrado para atualização em tempo real.", code=ErrCode.EXCEL_NOT_FOUND)

            try:
                with COMContext() as ctx:
                    if ctx is None or not ctx.has_workbooks:
                        return ToolResult.err("Excel aberto não encontrado para atualização em tempo real.", code=ErrCode.EXCEL_NOT_FOUND)
                    wb = ctx.resolve_workbook(path=workspace.path, name=workspace.excel_book_name)
                    if wb is None:
                        return ToolResult.err("Excel aberto não encontrado para atualização em tempo real.", code=ErrCode.EXCEL_NOT_FOUND)
                    ws = wb.Worksheets(workspace.sheet_name) if workspace.sheet_name else wb.ActiveSheet

                    for a in wb_actions:
                        kind = str(a.get("action", "")).strip().lower()
                        base = workspace.sheet_name or "Aba"
                        new_name = str(a.get("name", f"{base} Cópia")).strip() or f"{base} Cópia"
                        if kind == "duplicate_sheet":
                            src = wb.Worksheets(workspace.sheet_name) if workspace.sheet_name else wb.ActiveSheet
                            src.Copy(After=wb.Worksheets(wb.Worksheets.Count))
                            wb.Worksheets(wb.Worksheets.Count).Name = new_name
                        elif kind == "create_sheet":
                            wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count)).Name = new_name
                        elif kind == "delete_sheet":
                            target_name = str(a.get("name", "")).strip()
                            if target_name and wb.Worksheets.Count > 1:
                                excel_app = wb.Application
                                old_alerts = excel_app.DisplayAlerts
                                excel_app.DisplayAlerts = False
                                try:
                                    wb.Worksheets(target_name).Delete()
                                finally:
                                    excel_app.DisplayAlerts = old_alerts
                        elif kind == "rename_sheet":
                            old_name = str(a.get("from", "")).strip()
                            new_n = str(a.get("to", "")).strip()
                            if old_name and new_n:
                                wb.Worksheets(old_name).Name = new_n

                    if df_actions and new_df is not None:
                        ws.UsedRange.ClearContents()
                        if not new_df.empty:
                            nrows, ncols = new_df.shape[0] + 1, new_df.shape[1]
                            rows = _df_to_excel_matrix(new_df)
                            ws.Range(ws.Cells(1, 1), ws.Cells(nrows, ncols)).Value = rows
                        elif len(new_df.columns) > 0:
                            header = tuple(str(c) for c in new_df.columns)
                            ws.Range(ws.Cells(1, 1), ws.Cells(1, len(header))).Value = (header,)

                    wb.Save()
                    return ToolResult.ok(f"Otimização estruturada aplicada com sucesso (COM).{checkpoint_suffix}")
            except Exception as e:
                log.exception("Falha ao escrever no Excel via COM (structured_actions)")
                return ToolResult.err(f"Erro ao escrever no Excel: {e!s}", code=ErrCode.EXCEL_WRITE_COM)
        else:
            if wb_actions and new_wb is None:
                return ToolResult.err(
                    "Operações de aba (duplicar/criar/excluir/renomear) não são suportadas para arquivos .ods/.xls. "
                    "Salve o arquivo como .xlsx primeiro.",
                    code=ErrCode.FILE_INVALID_FORMAT,
                )
            if new_wb is not None and workspace.path:
                try:
                    for a in wb_actions:
                        kind = str(a.get("action", "")).strip().lower()
                        base = workspace.sheet_name or "Aba"
                        new_name = str(a.get("name", f"{base} Cópia")).strip() or f"{base} Cópia"
                        if kind == "duplicate_sheet":
                            src_ws = new_wb[workspace.sheet_name] if workspace.sheet_name and workspace.sheet_name in new_wb.sheetnames else new_wb.active
                            copy_ws = new_wb.copy_worksheet(src_ws)
                            copy_ws.title = new_name
                        elif kind == "create_sheet":
                            new_wb.create_sheet(title=new_name)
                        elif kind == "delete_sheet":
                            target_name = str(a.get("name", "")).strip()
                            if target_name and target_name in new_wb.sheetnames and len(new_wb.sheetnames) > 1:
                                del new_wb[target_name]
                        elif kind == "rename_sheet":
                            old_name = str(a.get("from", "")).strip()
                            new_n = str(a.get("to", "")).strip()
                            if old_name and new_n and old_name in new_wb.sheetnames:
                                new_wb[old_name].title = new_n

                    if df_actions and new_df is not None:
                        ws = new_wb[workspace.sheet_name] if workspace.sheet_name in new_wb.sheetnames else new_wb.active
                        ws.delete_rows(1, ws.max_row)
                        ws.append(list(new_df.columns))
                        for row in new_df.itertuples(index=False):
                            ws.append(list(row))
                    new_wb.save(workspace.path)
                except (OSError, PermissionError, ValueError) as e:
                    log.exception("Falha ao salvar arquivo com openpyxl (structured_actions)")
                    return ToolResult.err(f"Erro ao salvar arquivo: {e!s}", code=ErrCode.EXCEL_SAVE_OPENPYXL)
            elif df_actions and new_df is not None and workspace.path:
                try:
                    new_df.to_excel(workspace.path, index=False, engine="openpyxl")
                except (OSError, PermissionError, ValueError) as e:
                    log.exception("Falha ao salvar arquivo via to_excel (structured_actions)")
                    return ToolResult.err(f"Erro ao salvar arquivo: {e!s}", code=ErrCode.EXCEL_SAVE_DF)

        if df_actions and new_df is not None:
            workspace.df = new_df
            workspace.columns = list(new_df.columns)
            workspace.row_count = int(len(new_df))
            workspace.indexed_rows = int(len(new_df))
            workspace.truncated = False

        has_query = any(str(a.get("action", "")).strip().lower() in _QUERY_KINDS for a in df_actions)
        if has_query and new_df is not None:
            preview = new_df.to_string(max_rows=50)
            return ToolResult.ok(f"Resultado:\n{preview}{checkpoint_suffix}")

        return ToolResult.ok(f"Otimização estruturada aplicada com sucesso.{checkpoint_suffix}")
    finally:
        if wb_openpyxl is not None:
            wb_openpyxl.close()
