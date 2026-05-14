"""
Tools do agente:
- StructuredActions: aplica ações declarativas (sort, fillna, replace, etc.) sem exec()
- Sandbox: valida código (legado — usado apenas por testes de segurança)
"""
from __future__ import annotations

import ast
import json
import logging
from collections.abc import Callable
from pathlib import Path

from ..checkpoints.manager import CheckpointManager
from ..indexing.excel_reader import Workspace
from .result import ErrCode, ToolResult

log = logging.getLogger(__name__)

ALLOWED_MODULES = {"pandas", "openpyxl", "odf", "xlrd", "math", "datetime"}
FORBIDDEN_NAMES = {
    "__import__",
    "eval",
    "exec",
    "open",
    "compile",
    "globals",
    "locals",
    "input",
    "os",
    "sys",
    "subprocess",
    "getattr",
    "setattr",
    "delattr",
    "hasattr",
    "type",
    "vars",
    "dir",
    "breakpoint",
    "exit",
    "quit",
    "help",
    "memoryview",
    "bytearray",
    "classmethod",
    "staticmethod",
    "property",
    "super",
}
SAFE_BUILTINS = {
    "abs": abs,
    "all": all,
    "any": any,
    "bool": bool,
    "dict": dict,
    "enumerate": enumerate,
    "float": float,
    "int": int,
    "len": len,
    "list": list,
    "max": max,
    "min": min,
    "range": range,
    "round": round,
    "set": set,
    "sorted": sorted,
    "str": str,
    "sum": sum,
    "tuple": tuple,
    "zip": zip,
    "ValueError": ValueError,
    "TypeError": TypeError,
    "KeyError": KeyError,
    "IndexError": IndexError,
    "AttributeError": AttributeError,
    "RuntimeError": RuntimeError,
    "ZeroDivisionError": ZeroDivisionError,
    "StopIteration": StopIteration,
}


def _normalize_actions(actions_payload: str) -> tuple[list[dict], str]:
    """Aceita JSON em formato lista ou objeto {"actions": [...]}.
    Retorna (actions, erro)."""
    try:
        raw = json.loads(actions_payload)
    except Exception as e:
        log.warning("JSON inválido em [ACTIONS]: %s", e)
        return [], f"JSON inválido em [ACTIONS]: {e!s}"

    if isinstance(raw, dict):
        actions = raw.get("actions", [])
    else:
        actions = raw

    if not isinstance(actions, list) or not actions:
        return [], "[ACTIONS] precisa conter uma lista não vazia de ações."

    for i, a in enumerate(actions, 1):
        if not isinstance(a, dict):
            return [], f"Ação #{i} inválida: precisa ser objeto JSON."
        if not str(a.get("action", "")).strip():
            return [], f"Ação #{i} sem campo 'action'."
    return actions, ""


def _require_column(df, col: str) -> tuple[bool, str]:
    if col not in df.columns:
        return False, f"Coluna não encontrada: {col}"
    return True, ""


def _apply_actions_to_df(df, actions: list[dict]):
    """Aplica ações estruturadas (sem exec) sobre um DataFrame."""
    import pandas as pd

    out = df.copy()
    for idx, a in enumerate(actions, 1):
        kind = str(a.get("action", "")).strip().lower()

        if kind == "sort":
            by = a.get("by")
            if isinstance(by, str):
                by = [by]
            if not isinstance(by, list) or not by:
                return None, f"Ação #{idx} sort: campo 'by' inválido."
            for c in by:
                ok, err = _require_column(out, str(c))
                if not ok:
                    return None, f"Ação #{idx} sort: {err}"

            asc = a.get("ascending", True)
            na_pos = str(a.get("na_position", "last")).lower()
            if na_pos not in {"first", "last"}:
                na_pos = "last"
            out = out.sort_values(by=by, ascending=asc, na_position=na_pos, kind="mergesort").reset_index(drop=True)

        elif kind == "fillna":
            col = str(a.get("column", "")).strip()
            ok, err = _require_column(out, col)
            if not ok:
                return None, f"Ação #{idx} fillna: {err}"
            out[col] = out[col].fillna(a.get("value"))

        elif kind == "replace":
            col = str(a.get("column", "")).strip()
            ok, err = _require_column(out, col)
            if not ok:
                return None, f"Ação #{idx} replace: {err}"
            out[col] = out[col].replace(a.get("from"), a.get("to"))

        elif kind == "rename_column":
            old = str(a.get("from", "")).strip()
            new = str(a.get("to", "")).strip()
            if not old or not new:
                return None, f"Ação #{idx} rename_column: campos 'from' e 'to' são obrigatórios."
            ok, err = _require_column(out, old)
            if not ok:
                return None, f"Ação #{idx} rename_column: {err}"
            out = out.rename(columns={old: new})

        elif kind == "drop_columns":
            cols = a.get("columns")
            if isinstance(cols, str):
                cols = [cols]
            if not isinstance(cols, list) or not cols:
                return None, f"Ação #{idx} drop_columns: campo 'columns' inválido."
            missing = [c for c in cols if c not in out.columns]
            if missing:
                return None, f"Ação #{idx} drop_columns: colunas ausentes: {', '.join(map(str, missing))}"
            out = out.drop(columns=cols)

        elif kind == "add_computed_column":
            new_col = str(a.get("new_column", "")).strip()
            op = str(a.get("operation", "")).strip().lower()
            src = a.get("source_columns") or []
            if isinstance(src, str):
                src = [src]
            if not new_col or not isinstance(src, list) or not src:
                return None, f"Ação #{idx} add_computed_column: parâmetros inválidos."
            for c in src:
                ok, err = _require_column(out, str(c))
                if not ok:
                    return None, f"Ação #{idx} add_computed_column: {err}"

            if op == "concat":
                sep = str(a.get("separator", ""))
                out[new_col] = out[src].astype(str).agg(sep.join, axis=1)
            elif op in {"sum", "add"}:
                out[new_col] = out[src].apply(pd.to_numeric, errors="coerce").sum(axis=1)
            elif op == "multiply":
                out[new_col] = out[src].apply(pd.to_numeric, errors="coerce").prod(axis=1)
            elif op == "subtract":
                if len(src) != 2:
                    return None, f"Ação #{idx} add_computed_column/subtract: requer 2 colunas."
                nums = out[src].apply(pd.to_numeric, errors="coerce")
                out[new_col] = nums[src[0]] - nums[src[1]]
            elif op == "divide":
                if len(src) != 2:
                    return None, f"Ação #{idx} add_computed_column/divide: requer 2 colunas."
                nums = out[src].apply(pd.to_numeric, errors="coerce")
                denom = nums[src[1]].replace(0, pd.NA)
                out[new_col] = nums[src[0]] / denom
            else:
                return None, f"Ação #{idx} add_computed_column: operação não suportada: {op}"

        elif kind == "filter_equals":
            col = str(a.get("column", "")).strip()
            ok, err = _require_column(out, col)
            if not ok:
                return None, f"Ação #{idx} filter_equals: {err}"
            out = out[out[col] == a.get("value")].reset_index(drop=True)

        elif kind == "dropna":
            cols = a.get("columns")
            if cols is None:
                out = out.dropna(how=a.get("how", "any"))
            else:
                if isinstance(cols, str):
                    cols = [cols]
                if not isinstance(cols, list) or not cols:
                    return None, f"Ação #{idx} dropna: campo 'columns' inválido."
                missing = [c for c in cols if c not in out.columns]
                if missing:
                    return None, f"Ação #{idx} dropna: colunas ausentes: {', '.join(map(str, missing))}"
                how = str(a.get("how", "any")).lower()
                if how not in {"any", "all"}:
                    how = "any"
                out = out.dropna(subset=cols, how=how).reset_index(drop=True)

        elif kind == "groupby_agg":
            group_by = a.get("group_by")
            if isinstance(group_by, str):
                group_by = [group_by]
            if not isinstance(group_by, list) or not group_by:
                return None, f"Ação #{idx} groupby_agg: campo 'group_by' inválido."
            for c in group_by:
                ok, err = _require_column(out, str(c))
                if not ok:
                    return None, f"Ação #{idx} groupby_agg: {err}"
            agg_col = str(a.get("agg_column", "")).strip()
            ok, err = _require_column(out, agg_col)
            if not ok:
                return None, f"Ação #{idx} groupby_agg: {err}"
            agg_func = str(a.get("agg_func", "mean")).strip().lower()
            valid_funcs = {"mean", "sum", "min", "max", "count", "median", "std", "var", "first", "last"}
            if agg_func not in valid_funcs:
                return None, f"Ação #{idx} groupby_agg: função de agregação não suportada: {agg_func}. Use: {', '.join(sorted(valid_funcs))}"
            grouped = out.groupby(group_by, sort=False)[agg_col].agg(agg_func).reset_index()
            grouped.columns = list(group_by) + [agg_col]
            out = grouped

        elif kind == "filter_contains":
            col = str(a.get("column", "")).strip()
            ok, err = _require_column(out, col)
            if not ok:
                return None, f"Ação #{idx} filter_contains: {err}"
            value = str(a.get("value", ""))
            case = a.get("case_sensitive", False)
            if case:
                mask = out[col].astype(str).str.contains(value, regex=False, na=False)
            else:
                mask = out[col].astype(str).str.contains(value, regex=False, na=False, case=False)
            out = out[mask].reset_index(drop=True)

        elif kind == "filter_range":
            col = str(a.get("column", "")).strip()
            ok, err = _require_column(out, col)
            if not ok:
                return None, f"Ação #{idx} filter_range: {err}"
            nums = pd.to_numeric(out[col], errors="coerce")
            lo = a.get("min")
            hi = a.get("max")
            if lo is not None:
                nums = nums.where(nums >= float(lo))
            if hi is not None:
                nums = nums.where(nums <= float(hi))
            out = out[nums.notna()].reset_index(drop=True)

        elif kind == "pivot_table":
            index_cols = a.get("index")
            if isinstance(index_cols, str):
                index_cols = [index_cols]
            if not isinstance(index_cols, list) or not index_cols:
                return None, f"Ação #{idx} pivot_table: campo 'index' inválido (lista de colunas para agrupar)."
            for c in index_cols:
                ok, err = _require_column(out, str(c))
                if not ok:
                    return None, f"Ação #{idx} pivot_table: {err}"
            values_col = str(a.get("values", "")).strip()
            if values_col:
                ok, err = _require_column(out, values_col)
                if not ok:
                    return None, f"Ação #{idx} pivot_table: {err}"
            columns_col = a.get("columns")
            if isinstance(columns_col, str):
                columns_col = [columns_col]
            if columns_col:
                for c in columns_col:
                    if isinstance(c, str):
                        ok, err = _require_column(out, c)
                        if not ok:
                            return None, f"Ação #{idx} pivot_table: {err}"
            agg_func = str(a.get("agg_func", "sum")).strip().lower()
            valid_pv_funcs = {"mean", "sum", "min", "max", "count", "median", "std", "var", "first", "last"}
            if agg_func not in valid_pv_funcs:
                return None, f"Ação #{idx} pivot_table: agg_func não suportada: {agg_func}. Use: {', '.join(sorted(valid_pv_funcs))}"
            pv_kwargs: dict = {"index": index_cols, "aggfunc": agg_func}
            if values_col:
                pv_kwargs["values"] = values_col
            if columns_col:
                pv_kwargs["columns"] = columns_col
            try:
                pv = out.pivot_table(**pv_kwargs).reset_index()
            except Exception as e:
                return None, f"Ação #{idx} pivot_table: erro ao gerar pivot: {e!s}"
            out = pv

        elif kind == "merge_columns":
            cols = a.get("columns")
            if isinstance(cols, str):
                cols = [cols]
            if not isinstance(cols, list) or not cols:
                return None, f"Ação #{idx} merge_columns: campo 'columns' inválido."
            for c in cols:
                ok, err = _require_column(out, str(c))
                if not ok:
                    return None, f"Ação #{idx} merge_columns: {err}"
            new_col = str(a.get("new_column", "")).strip()
            if not new_col:
                return None, f"Ação #{idx} merge_columns: campo 'new_column' é obrigatório."
            sep = str(a.get("separator", " "))
            out[new_col] = out[cols].astype(str).agg(sep.join, axis=1)

        elif kind == "strip_whitespace":
            cols = a.get("columns")
            if cols is None:
                str_cols = [c for c in out.columns if pd.api.types.is_string_dtype(out[c])]
            else:
                if isinstance(cols, str):
                    cols = [cols]
                if not isinstance(cols, list) or not cols:
                    return None, f"Ação #{idx} strip_whitespace: campo 'columns' inválido."
                str_cols = [str(c) for c in cols]
                for c in str_cols:
                    ok, err = _require_column(out, c)
                    if not ok:
                        return None, f"Ação #{idx} strip_whitespace: {err}"
            for c in str_cols:
                if c in out.columns and pd.api.types.is_string_dtype(out[c]):
                    out[c] = out[c].astype(str).str.strip()

        elif kind == "change_dtype":
            col = str(a.get("column", "")).strip()
            ok, err = _require_column(out, col)
            if not ok:
                return None, f"Ação #{idx} change_dtype: {err}"
            dtype = str(a.get("dtype", "")).strip().lower()
            valid_dtypes = {"int", "float", "str", "bool", "datetime"}
            if dtype not in valid_dtypes:
                return None, f"Ação #{idx} change_dtype: tipo não suportado: {dtype}. Use: {', '.join(sorted(valid_dtypes))}"
            try:
                if dtype == "int":
                    out[col] = pd.to_numeric(out[col], errors="coerce").astype("Int64")
                elif dtype == "float":
                    out[col] = pd.to_numeric(out[col], errors="coerce").astype(float)
                elif dtype == "str":
                    out[col] = out[col].astype(str)
                elif dtype == "bool":
                    out[col] = out[col].astype(bool)
                elif dtype == "datetime":
                    out[col] = pd.to_datetime(out[col], errors="coerce")
            except Exception as e:
                return None, f"Ação #{idx} change_dtype: erro ao converter coluna '{col}' para {dtype}: {e!s}"

        elif kind == "adjust_header":
            from ..indexing.excel_reader import _detect_header_offset
            offset = _detect_header_offset(out)
            if offset <= 0:
                return None, f"Ação #{idx} adjust_header: não foi possível detectar um cabeçalho alternativo."
            return out, f"__ADJUST_HEADER__{offset}"

        elif kind == "request_more_rows":
            return out, "__REQUEST_MORE_ROWS__"

        else:
            return None, f"Ação #{idx} desconhecida: {kind}"

    return out, ""


def _validate_code(code: str) -> tuple[bool, str]:
    """Verifica sintaxe e imports permitidos."""
    try:
        tree = ast.parse(code)
    except SyntaxError as e:
        return False, f"Erro de sintaxe: {e}"

    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                if alias.name.split(".")[0] not in ALLOWED_MODULES:
                    return False, f"Módulo não permitido: {alias.name}"
        if isinstance(node, ast.ImportFrom):
            if node.module and node.module.split(".")[0] not in ALLOWED_MODULES:
                return False, f"Módulo não permitido: {node.module}"
        if isinstance(node, ast.Name) and node.id in FORBIDDEN_NAMES:
            return False, f"Nome não permitido: {node.id}"
        if isinstance(node, ast.Attribute):
            attr = str(node.attr)
            if attr.startswith("__"):
                return False, "Acesso a atributos internos (__dunder__) não é permitido."
            if attr in FORBIDDEN_NAMES:
                return False, f"Acesso a atributo não permitido: {attr}"
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name) and node.func.id in FORBIDDEN_NAMES:
            return False, f"Função não permitida: {node.func.id}"
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Attribute):
            if node.func.attr in FORBIDDEN_NAMES:
                return False, f"Chamada de método não permitida: {node.func.attr}"

    return True, ""


def _excel_scalar(value):
    """Converte valores pandas/numpy para tipos seguros para COM do Excel."""
    try:
        import pandas as pd  # type: ignore
        if pd.isna(value):
            return None
    except Exception:
        log.exception("Falha ao verificar valor NA com pandas")

    try:
        import numpy as np  # type: ignore
        if isinstance(value, np.generic):
            return value.item()
    except Exception:
        log.exception("Falha ao converter valor escalar numpy")

    return value


def _df_to_excel_matrix(df):
    rows = [tuple(str(c) for c in df.columns)]
    for row in df.itertuples(index=False, name=None):
        rows.append(tuple(_excel_scalar(v) for v in row))
    return tuple(rows)


def _classify_err(msg: str) -> ErrCode:
    if "Coluna" in msg and ("não encontrada" in msg or "ausente" in msg):
        return ErrCode.COLUMN_MISSING
    if "desconhecida" in msg:
        return ErrCode.ACTION_UNKNOWN
    return ErrCode.ACTION_INVALID


_QUERY_KINDS = {"groupby_agg", "pivot_table"}

_WB_KINDS = {"duplicate_sheet", "create_sheet", "delete_sheet", "rename_sheet"}


def structured_actions_tool(
    workspace: Workspace,
    actions_payload: str,
    checkpoint_manager: CheckpointManager,
    on_checkpoint_saved: Callable[[str], None] | None = None,
    save_checkpoint: bool = True,
) -> ToolResult:
    """
    Executa ações estruturadas (JSON) sobre a planilha sem usar exec arbitrário.
    """
    actions, parse_err = _normalize_actions(actions_payload)
    if parse_err:
        return ToolResult.err(parse_err, code=ErrCode.PARSE_ACTIONS)

    checkpoint_suffix = ""
    if save_checkpoint:
        try:
            info = checkpoint_manager.save_checkpoint(workspace, label="Antes da ordem")
            if on_checkpoint_saved:
                on_checkpoint_saved(info.label)
            checkpoint_suffix = " Checkpoint salvo antes da ordem."
        except Exception as cp_err:
            log.exception("Falha ao salvar checkpoint (structured_actions)")
            return ToolResult.err(f"Erro ao salvar checkpoint: {cp_err!s}", code=ErrCode.CHECKPOINT_SAVE)

    import openpyxl
    import pandas as pd

    df = workspace.df
    if df is None:
        df = pd.DataFrame(columns=workspace.columns or [])

    # Separate workbook-level actions (sheet ops) from df-level actions
    wb_actions = [a for a in actions if str(a.get("action", "")).strip().lower() in _WB_KINDS]
    df_actions = [a for a in actions if str(a.get("action", "")).strip().lower() not in _WB_KINDS]

    if df_actions:
        new_df, err = _apply_actions_to_df(df, df_actions)
        if err:
            if err.startswith("__REQUEST_MORE_ROWS__") or err.startswith("__ADJUST_HEADER__"):
                return ToolResult.ok(err)
            return ToolResult.err(err, code=_classify_err(err))
    else:
        new_df = df.copy() if df is not None else pd.DataFrame(columns=workspace.columns or [])

    wb_openpyxl = None
    if workspace.path and not workspace.excel_live:
        save_path = workspace.path
        original_suffix = Path(save_path).suffix.lower()
        if original_suffix in (".ods", ".xls"):
            save_path = str(Path(save_path).with_suffix(".xlsx"))
            workspace.path = save_path
            log.info("Conversão de formato: %s → %s", original_suffix, ".xlsx")
        try:
            wb_openpyxl = openpyxl.load_workbook(save_path)
        except Exception:
            log.exception("Falha ao carregar workbook com openpyxl (structured_actions)")
            wb_openpyxl = None

    new_wb = wb_openpyxl
    if workspace.excel_live:
        try:
            from ..com_utils import COMContext

            ctx = COMContext()
            ctx.__enter__()
        except Exception:
            ctx = None
        try:
            if ctx is None or not ctx.has_workbooks:
                return ToolResult.err("Excel aberto não encontrado para atualização em tempo real.", code=ErrCode.EXCEL_NOT_FOUND)
            wb = ctx.resolve_workbook(path=workspace.path, name=workspace.excel_book_name)
            if wb is None:
                return ToolResult.err("Excel aberto não encontrado para atualização em tempo real.", code=ErrCode.EXCEL_NOT_FOUND)
            ws = wb.Worksheets(workspace.sheet_name) if workspace.sheet_name else wb.ActiveSheet

            # Apply workbook-level actions (COM)
            for a in wb_actions:
                kind = str(a.get("action", "")).strip().lower()
                base = workspace.sheet_name or "Aba"
                new_name = str(a.get("name", f"{base} Cópia")).strip() or f"{base} Cópia"
                if kind == "duplicate_sheet":
                    src = wb.Worksheets(workspace.sheet_name) if workspace.sheet_name else wb.ActiveSheet
                    src.Copy(After=wb.Worksheets(wb.Worksheets.Count))
                    wb.Worksheets(wb.Worksheets.Count).Name = new_name
                elif kind == "create_sheet":
                    wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count)).Name = new_name
                elif kind == "delete_sheet":
                    target_name = str(a.get("name", "")).strip()
                    if target_name and wb.Worksheets.Count > 1:
                        excel_app = wb.Application
                        old_alerts = excel_app.DisplayAlerts
                        excel_app.DisplayAlerts = False
                        try:
                            wb.Worksheets(target_name).Delete()
                        finally:
                            excel_app.DisplayAlerts = old_alerts
                elif kind == "rename_sheet":
                    old_name = str(a.get("from", "")).strip()
                    new_n = str(a.get("to", "")).strip()
                    if old_name and new_n:
                        wb.Worksheets(old_name).Name = new_n

            if df_actions and new_df is not None:
                ws.UsedRange.ClearContents()
                if not new_df.empty:
                    nrows, ncols = new_df.shape[0] + 1, new_df.shape[1]
                    rows = _df_to_excel_matrix(new_df)
                    ws.Range(ws.Cells(1, 1), ws.Cells(nrows, ncols)).Value = rows
                elif len(new_df.columns) > 0:
                    header = tuple(str(c) for c in new_df.columns)
                    ws.Range(ws.Cells(1, 1), ws.Cells(1, len(header))).Value = (header,)

            wb.Save()
            return ToolResult.ok(f"Otimização estruturada aplicada com sucesso (COM).{checkpoint_suffix}")
        except Exception as e:
            log.exception("Falha ao escrever no Excel via COM (structured_actions)")
            return ToolResult.err(f"Erro ao escrever no Excel: {e!s}", code=ErrCode.EXCEL_WRITE_COM)
        finally:
            if ctx is not None:
                ctx.__exit__(None, None, None)
    else:
        if wb_actions and new_wb is None:
            return ToolResult.err(
                "Operações de aba (duplicar/criar/excluir/renomear) não são suportadas para arquivos .ods/.xls. "
                "Salve o arquivo como .xlsx primeiro.",
                code=ErrCode.FILE_INVALID_FORMAT,
            )
        if new_wb is not None and workspace.path:
            try:
                for a in wb_actions:
                    kind = str(a.get("action", "")).strip().lower()
                    base = workspace.sheet_name or "Aba"
                    new_name = str(a.get("name", f"{base} Cópia")).strip() or f"{base} Cópia"
                    if kind == "duplicate_sheet":
                        src_ws = new_wb[workspace.sheet_name] if workspace.sheet_name and workspace.sheet_name in new_wb.sheetnames else new_wb.active
                        copy_ws = new_wb.copy_worksheet(src_ws)
                        copy_ws.title = new_name
                    elif kind == "create_sheet":
                        new_wb.create_sheet(title=new_name)
                    elif kind == "delete_sheet":
                        target_name = str(a.get("name", "")).strip()
                        if target_name and target_name in new_wb.sheetnames and len(new_wb.sheetnames) > 1:
                            del new_wb[target_name]
                    elif kind == "rename_sheet":
                        old_name = str(a.get("from", "")).strip()
                        new_n = str(a.get("to", "")).strip()
                        if old_name and new_n and old_name in new_wb.sheetnames:
                            new_wb[old_name].title = new_n

                if df_actions and new_df is not None:
                    ws = new_wb[workspace.sheet_name] if workspace.sheet_name in new_wb.sheetnames else new_wb.active
                    ws.delete_rows(1, ws.max_row)
                    ws.append(list(new_df.columns))
                    for row in new_df.itertuples(index=False):
                        ws.append(list(row))
                new_wb.save(workspace.path)
            except Exception as e:
                log.exception("Falha ao salvar arquivo com openpyxl (structured_actions)")
                return ToolResult.err(f"Erro ao salvar arquivo: {e!s}", code=ErrCode.EXCEL_SAVE_OPENPYXL)
        elif df_actions and new_df is not None and workspace.path:
            try:
                new_df.to_excel(workspace.path, index=False, engine="openpyxl")
            except Exception as e:
                log.exception("Falha ao salvar arquivo via to_excel (structured_actions)")
                return ToolResult.err(f"Erro ao salvar arquivo: {e!s}", code=ErrCode.EXCEL_SAVE_DF)

    if df_actions and new_df is not None:
        workspace.df = new_df
        workspace.columns = list(new_df.columns)
        workspace.row_count = int(len(new_df))
        workspace.indexed_rows = int(len(new_df))
        workspace.truncated = False

    has_query = any(str(a.get("action", "")).strip().lower() in _QUERY_KINDS for a in df_actions)
    if has_query and new_df is not None:
        preview = new_df.to_string(max_rows=50)
        return ToolResult.ok(f"Resultado:\n{preview}{checkpoint_suffix}")

    return ToolResult.ok(f"Otimização estruturada aplicada com sucesso.{checkpoint_suffix}")
