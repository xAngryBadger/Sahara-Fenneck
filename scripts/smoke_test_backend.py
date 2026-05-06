from pathlib import Path
import sys
import tempfile
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.indexing.excel_reader import index_from_path
from src.checkpoints.manager import CheckpointManager
from src.agent.tools import optimize_tool
from src.agent.runner import run_agent


def assert_true(cond, msg):
    if not cond:
        raise AssertionError(msg)


def create_sample_xlsx(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendas"
    ws.append(["Vendas", "Regiao"])
    ws.append([100, "Sul"])
    ws.append([200, "Norte"])
    wb.save(path)


def read_headers(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    first_data = [c.value for c in ws[2]]
    wb.close()
    return headers, first_data


def main():
    with tempfile.TemporaryDirectory() as tmp:
        xlsx = Path(tmp) / "teste.xlsx"
        create_sample_xlsx(xlsx)

        # 1) Indexacao
        ws = index_from_path(str(xlsx))
        assert_true(ws.error is None, f"indexacao falhou: {ws.error}")
        assert_true(ws.row_count == 2, f"row_count inesperado: {ws.row_count}")

        # 2) Optimize + checkpoint antes da alteracao
        cp = CheckpointManager(str(xlsx), interaction_label="Teste")
        code = "df['Vendas_Dobro'] = df['Vendas'] * 2"
        result = optimize_tool(ws, code, cp)
        assert_true("sucesso" in result.lower(), f"optimize falhou: {result}")

        # 3) Verificar alteracao aplicada no arquivo
        headers, _ = read_headers(xlsx)
        assert_true("Vendas_Dobro" in headers, "coluna nova nao foi salva no arquivo")

        # 4) Verificar checkpoints e restauracao
        cps = cp.list_checkpoints()
        assert_true(len(cps) >= 1, "checkpoint nao foi criado")
        ok = cp.restore(cps[-1].path)
        assert_true(ok, "falha ao restaurar checkpoint")

        # 5) Confirmar restore removendo coluna adicionada
        headers_after_restore, _ = read_headers(xlsx)
        assert_true("Vendas_Dobro" not in headers_after_restore, "restore nao voltou ao estado anterior")

        # 6) run_agent sem Ollama (se nao houver) deve responder com mensagem clara
        msg = run_agent("some a coluna vendas", ws)
        assert_true(isinstance(msg, str) and len(msg) > 0, "run_agent retornou vazio")

    print("SMOKE_TEST_OK")


if __name__ == "__main__":
    main()
