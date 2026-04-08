from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def gerar_tabela_hora_homem(
    xlsx_path: Path,
    itens: list[tuple[str, str, float]],
    colaboradores_lista=None,
    jornada_por_colaborador=4.3,
    coluna_inicio=15,  # 15 = coluna O (1=A)
):
    if colaboradores_lista is None:
        colaboradores_lista = [1, 2, 3, 4, 5, 6, 7, 8]

    if not xlsx_path.exists():
        # cria arquivo base com sheet Planilha1
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Planilha1"
        wb.save(xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.title = "Planilha1"

    headers = [
        "Área",
        "Atividade",
        "HH_por_hectare",
        "Colaboradores",
        "HH_dia_total",
        "Hectares_por_dia",
        "Dias_para_1_ha",
        "Horas_sobra_ao_1ha",
        "Observação",
    ]

    max_col_cleanup = coluna_inicio + len(headers) + 2
    for row in range(1, ws.max_row + 1):
        for col in range(coluna_inicio, max_col_cleanup + 1):
            ws.cell(row=row, column=col).value = None

    for idx, header in enumerate(headers):
        ws.cell(row=1, column=coluna_inicio + idx, value=header)

    linha = 2
    for area, atividade, hh_ha in itens:
        for n_colab in colaboradores_lista:
            hh_dia = round(n_colab * jornada_por_colaborador, 4)
            hectares_dia = round(hh_dia / hh_ha if hh_ha else 0, 4)
            dias_1ha = round(hh_ha / hh_dia if hh_dia else float("inf"), 4)
            sobra = round(hh_dia - hh_ha, 4) if hh_dia > hh_ha else 0.0
            obs = (
                f"1ha em {dias_1ha}d, sobra {sobra} hh"
                if hh_dia >= hh_ha
                else f"falta {round(hh_ha - hh_dia,4)} hh"
            )

            values = [
                area,
                atividade,
                hh_ha,
                n_colab,
                hh_dia,
                hectares_dia,
                dias_1ha,
                sobra,
                obs,
            ]

            for idx, value in enumerate(values):
                ws.cell(row=linha, column=coluna_inicio + idx, value=value)
            linha += 1

    for idx in range(len(headers)):
        col_letter = get_column_letter(coluna_inicio + idx)
        ws.column_dimensions[col_letter].width = 16

    wb.save(xlsx_path)
    return xlsx_path


if __name__ == "__main__":
    arquivo = Path("E:/Sahara Fenneck/tabela papai.xlsx")
    entradas = [
        ("ri2ac0010", "ADUBAÇÃO QUÍM MAN DE BASE", 0.107),
        ("ri2ac0010", "CAPINA MANUAL COROA", 0.107),
        ("ri2ac0013", "ROÇADA MANUAL", 0.244),
        ("ri2ac0013", "IRRIGAÇÃO INICIAL", 0.244),
    ]
    resultado = gerar_tabela_hora_homem(arquivo, entradas, colaboradores_lista=[1,2,3,4,5,6,7,8], coluna_inicio=15)
    print("Tabela gerada em", resultado)
