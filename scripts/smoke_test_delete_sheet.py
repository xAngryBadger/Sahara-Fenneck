# -*- coding: utf-8 -*-
"""Smoke test: delete_sheet and rename_sheet actions (file-based)."""
import sys, pathlib, tempfile, shutil

ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
from src.indexing.excel_reader import index_file_multi
from src.checkpoints.manager import CheckpointManager
from src.agent.tools import structured_actions_tool

# 1. Create a temp workbook with 3 sheets
tmp = pathlib.Path(tempfile.mkdtemp()) / "test_sheets.xlsx"
wb = openpyxl.Workbook()
wb.active.title = "Dados"
wb.active.append(["Col1", "Col2"])
wb.active.append([1, 2])
wb.create_sheet("Lixo1").append(["A"])
wb.create_sheet("Lixo2").append(["B"])
wb.save(str(tmp))
wb.close()

# 2. Index the workbook
ws_list = index_file_multi(str(tmp), include_all_sheets=True, max_rows=0)
ws = [w for w in ws_list if w.sheet_name == "Dados"][0]
cp = CheckpointManager(str(tmp), interaction_label="Test")

# 3. Delete Lixo1 via structured actions
payload = '{"actions": [{"action": "delete_sheet", "name": "Lixo1"}]}'
result = structured_actions_tool(ws, payload, cp, save_checkpoint=False)
assert "sucesso" in result.lower(), f"Expected success: {result}"

# Verify Lixo1 is gone
wb2 = openpyxl.load_workbook(str(tmp))
assert "Lixo1" not in wb2.sheetnames, f"Lixo1 should be deleted: {wb2.sheetnames}"
assert "Dados" in wb2.sheetnames
assert "Lixo2" in wb2.sheetnames
wb2.close()

# 4. Rename Lixo2 to Backup via structured actions
payload2 = '{"actions": [{"action": "rename_sheet", "from": "Lixo2", "to": "Backup"}]}'
result2 = structured_actions_tool(ws, payload2, cp, save_checkpoint=False)
assert "sucesso" in result2.lower(), f"Expected success: {result2}"

wb3 = openpyxl.load_workbook(str(tmp))
assert "Backup" in wb3.sheetnames, f"Lixo2 should be renamed: {wb3.sheetnames}"
assert "Lixo2" not in wb3.sheetnames
wb3.close()

# Cleanup
shutil.rmtree(tmp.parent, ignore_errors=True)

print("SMOKE_DELETE_RENAME_SHEET_OK")
